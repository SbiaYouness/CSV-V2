import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


_COMPARISON_STATUSES = {
    "OK": "matched",
    "Non trouvé dans le PDF": "pdf_only",
    "Absent du fichier resultats": "csv_only",
    "ECART SIGNIFICATIF": "mismatched",
    "ANOMALIE UNITE PROBABLE (facteur ~100)": "mismatched",
    "ANOMALIE UNITE PROBABLE (facteur ~100, fichier resultats)": "mismatched",
    "PDF non disponible (pas de LEI)": "pdf_only",
}

_REF_ALIASES = ["reference", "référence", "libelle", "libellé", "description", "libellé/référence"]
_DATE_ALIASES = ["date", "date de l'opération", "valeur"]


def _nk(s: str) -> str:
    """Accent-strip + casefold a string for collision-safe dict lookup.
    Eliminates the entire class of mojibake/accent-encoding mismatches.
    """
    nfkd = unicodedata.normalize("NFKD", str(s).strip())
    ascii_s = nfkd.encode("ascii", "ignore").decode("ascii")
    return ascii_s.lower()


# ─── LEI lookup: keyed on accent-normalised column name ──────────────────────
# LEIs are permanent ISO 17442 identifiers — these never change.
# Banks without a corresponding ZIP in FichiersClaude/pdfs/ are listed with
# their real LEI so the lookup works as soon as a ZIP is added.
_BANK_LEI_MAP_RAW: dict[str, str] = {
    # Banks WITH ZIPs currently available
    "Groupe BPCE":                        "FR9695005MSX1OYEMGDF",
    "BANQUE POPULAIRE AUVERGNE RHON":     "969500JM7VIGQIPZOL49",
    "BNP Paribas":                        "R0MUWSFPU8MPRO8K5P83",
    "BRED - Banque populaire":            "NICH5Q04ADUV9SN3Q390",
    "Banque populaire Alsace Lorrai":     "969500EVOBAGHKZEXA33",
    "Banque populaire Rives de Pari":     "969500W8SBCXNX1DG443",
    "Bpifrance":                          "969500STN7T9MRUMJ267",
    "CAISSE D EPARGNE ET DE PREVOYA":     "969500JJWO4PQG0R1C58",  # Aquitaine Poitou-Charentes
    "Caisse d Epargne CEPAC":             "969500TU5ZMYBIWP0R51",
    # accent variants
    "Caisse d épargne et de prévoya":     "NO_LEI_EMPTY_COL",  # Mostly empty column in Excel
    "Caisse d épargne et de prévoya.1":  "969500SJXM8MW32ZVG75",  # Bretagne
    "Caisse d épargne et de prévoya.2":  "969500W34S6NCZWYBV47",  # Ile-de-France
    "Caisse d épargne et de prévoya.3":  "969500VR2NA6ANMTXH21",  # Rhone Alpes
    "Credit Foncier":                     "969500EYG6U339D3TI84",
    "Credit mutuel Arkea":                "96950041VJ1QP0B69503",
    "Natixis":                            "KX1WK48MPD4Y2NCUIZ63",
    "SFIL S A":                           "549300HFEHJOXGE4ZE63",
    "Societe generale S A":               "O2RNE8IBXP4R0TD8PU41",
    # Banks currently without ZIPs — real LEIs for future use
    "AGENCE FRANCE LOCALE - SOCIETE":    "969500QDXBKBTYXRS977",
    "Axa banque":                         "969500FFKE8NKHXUTZ47",
    "BNP Paribas Personal Finance":       "R0MUWSFPU8MPRO8K5P83",  # same group, placeholder
    "Banque populaire Grand Ouest":       "969500JNCCPQSAQO9K33",
    "BofA Securities Europe":             "R4DHLH5KX7NMDO6YTPM4",
    "ABN AMRO Bank N.V.":                 "BFXS5XCH7N0Y05NIXW11",
    "CIC":                                "96950066NYMAFZAJFW89",
    "Caisse federale de credit mutuel":   "969500RIQZX7RZNR4Z50",
    "Caisse regionale de credit agr":     "9695001XNQTXW5QYMM16",
    "Caisse regionale de credit agr.1":  "969500LP6CBLEFQ23P13",
    "Confederation Nationale du Cr":      "969500DZUEMFZXPMKV94",
    "RCI Banque":                         "969500V535YV4YSOVQ94",
}
# Build a normalised-key lookup dict once at import time
_BANK_LEI_MAP: dict[str, str] = {_nk(k): v for k, v in _BANK_LEI_MAP_RAW.items()}


def _lei_for_col(col: str) -> str:
    """Return LEI for an Excel column name, accent-insensitive."""
    if col.strip() in _BANK_LEI_MAP_RAW:
        return _BANK_LEI_MAP_RAW[col.strip()]
        
    nk_col = _nk(col)
    if nk_col in _BANK_LEI_MAP:
        return _BANK_LEI_MAP[nk_col]
    
    # Prefix / Substring lookup
    for k, v in _BANK_LEI_MAP.items():
        if nk_col.startswith(k) or k.startswith(nk_col):
            return v
    return ""


# ─── Output bank name map (accent-safe keys) ─────────────────────────────────
_OUTPUT_BANK_NAME_MAP_RAW: dict[str, str] = {
    "Groupe BPCE":                        "Groupe BPCE",
    "AGENCE FRANCE LOCALE - SOCIETE":    "Agence France Locale",
    "Axa banque":                         "Axa banque",
    "BANQUE POPULAIRE AUVERGNE RHON":     "Banque Populaire Auvergne Rhône Alpes",
    "BNP Paribas":                        "BNP Paribas",
    "BNP Paribas Personal Finance":       "BNP Paribas Personal Finance",
    "BRED - Banque populaire":            "BRED - Banque populaire",
    "Banque populaire Alsace Lorrai":     "Banque populaire Alsace Lorraine Champagne",
    "Banque populaire Grand Ouest":       "Banque populaire Grand Ouest",
    "Banque populaire Rives de Pari":     "Banque populaire Rives de Paris",
    "BofA Securities Europe":             "BofA Securities Europe",
    "ABN AMRO Bank N.V.":                 "ABN AMRO Bank N.V.",
    "Bpifrance":                          "Bpifrance",
    "CAISSE D EPARGNE ET DE PREVOYA":     "Caisse d'épargne Aquitaine Poitou-Charentes",
    "CIC":                                "CIC",
    "Caisse d Epargne CEPAC":             "Caisse d'Epargne CEPAC",
    "Caisse d épargne et de prévoya":     "Caisse d'épargne (Autre/Vide)",
    "Caisse d épargne et de prévoya.1":  "Caisse d'épargne Bretagne-Pays de Loire",
    "Caisse d épargne et de prévoya.2":  "Caisse d'épargne Ile-de-France",
    "Caisse d épargne et de prévoya.3":  "Caisse d'épargne Rhône Alpes",
    "Caisse federale de credit mutu":     "Caisse fédérale de crédit mutuel",
    "Caisse federale de credit mutuel":   "Caisse fédérale de crédit mutuel",
    "Caisse regionale de credit agr.1":  "Caisse régionale de crédit agricole (1)",
    "Caisse regionale de credit agr":    "Caisse régionale de crédit agricole (2)",
    "Confederation Nationale du Cr":      "Confédération Nationale du Crédit Mutuel",
    "Confederation Nationale du Cre":     "Confédération Nationale du Crédit Mutuel",
    "Credit Foncier":                     "Credit Foncier de France",
    "Credit mutuel Arkea":                "Crédit mutuel Arkéa",
    "Natixis":                            "Natixis",
    "RCI Banque":                         "RCI Banque",
    "SFIL S A":                           "SFIL S.A.",
    "Societe generale S A":               "Société générale S.A.",
}
_OUTPUT_BANK_NAME_MAP: dict[str, str] = {_nk(k): v for k, v in _OUTPUT_BANK_NAME_MAP_RAW.items()}


def _output_name_for_col(col: str) -> str:
    """Return the clean output bank name for an Excel column, accent-insensitive."""
    if col.strip() in _OUTPUT_BANK_NAME_MAP_RAW:
        return _OUTPUT_BANK_NAME_MAP_RAW[col.strip()]

    nk_col = _nk(col)
    if nk_col in _OUTPUT_BANK_NAME_MAP:
        return _OUTPUT_BANK_NAME_MAP[nk_col]
        
    for k, v in _OUTPUT_BANK_NAME_MAP.items():
        if nk_col.startswith(k) or k.startswith(nk_col):
            return v
    return str(col).strip()

_EXPORT_COLUMNS = [
    "Entité",
    "LEI",
    "Indicateur",
    "Valeur resultats",
    "Valeur PDF (EBA)",
    "Ecart",
    "Ecart %",
    "Statut",
    "Source PDF",
]


def _normalize_key(value: str) -> str:
    return str(value).strip().lower().replace("é", "e").replace("è", "e").replace("ê", "e").replace("à", "a").replace("ç", "c")


def _pick_bank_column(df: pd.DataFrame, pdf_name: str | None = None) -> str | None:
    if len(df.columns) <= 1:
        return None

    candidate_columns = [col for col in df.columns[1:] if str(col).strip()]
    if not candidate_columns:
        return None

    if pdf_name:
        pdf_key = _normalize_key(pdf_name)
        
        # 1. Try matching by LEI embedded in the filename
        # pdf_name usually contains the LEI at the beginning.
        for col in candidate_columns:
            lei = _lei_for_col(col)
            if lei and lei.lower() in pdf_key:
                return col

        # 2. Try matching by name substring
        if "bpce" in pdf_key:
            for col in candidate_columns:
                if "bpce" in _normalize_key(col):
                    return col
        for col in candidate_columns:
            if _normalize_key(col) in pdf_key or pdf_key in _normalize_key(col):
                return col

    # Fallback: pick the column with the most data
    best_column = None
    best_score = -1
    for col in candidate_columns:
        score = int(df[col].notna().sum())
        if score > best_score:
            best_score = score
            best_column = col
    return best_column


def _find_column(df: pd.DataFrame, aliases: list[str]) -> str | None:
    col_map = {col.lower().strip(): col for col in df.columns}
    for alias in aliases:
        if alias in col_map:
            return col_map[alias]
    return None


def _parse_amount(value) -> float:
    if pd.isna(value) or str(value).strip() == "":
        return 0.0
    
    raw = str(value).strip().replace("\u00a0", "").replace(" ", "")
    if not raw:
        return 0.0
        
    if "," in raw and "." in raw:
        if raw.rfind(",") > raw.rfind("."):
            raw = raw.replace(".", "").replace(",", ".")
        else:
            raw = raw.replace(",", "")
    elif "," in raw:
        raw = raw.replace(",", ".")
        
    try:
        return float(raw)
    except ValueError:
        return 0.0


def _to_export_rows(result: dict) -> list[dict]:
    rows = result.get("details") or result.get("transactions") or []
    export_rows: list[dict] = []

    for row in rows:
        export_rows.append({
            "Entité": row.get("Entité") or row.get("entity") or "Groupe BPCE",
            "LEI": row.get("LEI") or row.get("lei") or "",
            "Indicateur": row.get("Indicateur") or row.get("indicator") or row.get("reference") or "",
            "Valeur resultats": row.get("Valeur resultats") if row.get("Valeur resultats") is not None else row.get("csv_amount") if row.get("csv_amount") is not None else row.get("amount"),
            "Valeur PDF (EBA)": row.get("Valeur PDF (EBA)") if row.get("Valeur PDF (EBA)") is not None else row.get("pdf_amount") if row.get("pdf_amount") is not None else row.get("pdfValue"),
            "Ecart": row.get("Ecart") if row.get("Ecart") is not None else row.get("difference"),
            "Ecart %": row.get("Ecart %") if row.get("Ecart %") is not None else row.get("difference_pct"),
            "Statut": row.get("Statut") or row.get("status") or "",
            "Source PDF": row.get("Source PDF") or row.get("sourcePdf") or row.get("date") or "",
        })

    return export_rows


def build_comparison_workbook(result: dict, output_path: str | Path) -> str:
    """Write the comparison result to an Excel workbook."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Comparaison"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    worksheet.append(_EXPORT_COLUMNS)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in _to_export_rows(result):
        worksheet.append([row.get(column, "") for column in _EXPORT_COLUMNS])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    summary = workbook.create_sheet("Résumé")
    summary_rows = [
        ("Score de conformité", result.get("score", 0.0)),
        ("Correspondances", result.get("matched", 0)),
        ("Ecarts", result.get("mismatched", 0)),
        ("PDF uniquement", result.get("pdf_only", 0)),
        ("Resultats uniquement", result.get("csv_only", 0)),
        ("Total", len(_to_export_rows(result))),
        ("Synthèse IA", result.get("aiSynthesis", "")),
    ]
    for label, value in summary_rows:
        summary.append([label, value])

    workbook.save(output_path)
    return str(output_path)


def build_batch_comparison_workbook(results: list[dict], output_path: str | Path) -> str:
    """Write multiple comparison results into a single workbook."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    for index, result in enumerate(results, start=1):
        pdf_name = str(result.get("pdf_name") or result.get("pdfFile", {}).get("name") or f"PDF {index}")
        sheet_name = pdf_name.replace("[", "(").replace("]", ")")[:31] or f"PDF {index}"
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.append(_EXPORT_COLUMNS)
        for row in _to_export_rows(result):
            worksheet.append([row.get(column, "") for column in _EXPORT_COLUMNS])
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

    workbook.save(output_path)
    return str(output_path)


def compare_transactions(
    pdf_transactions: list[dict],
    csv_df: pd.DataFrame,
    pdf_name: str | None = None,
) -> dict:
    comparison_columns = {"Entité", "Indicateur", "Valeur resultats"}
    if comparison_columns.issubset(set(csv_df.columns)):
        rows = []
        matched = 0
        mismatched = 0
        pdf_only = 0
        csv_only = 0

        for _, row in csv_df.iterrows():
            indicator = str(row.get("Indicateur", "")).strip()
            result_value = _parse_amount(row.get("Valeur resultats"))
            pdf_value = _parse_amount(row.get("Valeur PDF (EBA)") if "Valeur PDF (EBA)" in csv_df.columns else None)
            status = str(row.get("Statut", "")).strip()

            if not status:
                if pd.isna(result_value) and pd.isna(pdf_value):
                    status = "PDF non disponible (pas de LEI)"
                elif pd.isna(pdf_value):
                    status = "Non trouvé dans le PDF"
                elif pd.isna(result_value):
                    status = "Absent du fichier resultats"
                else:
                    delta = abs(result_value - pdf_value)
                    delta_pct = (delta / abs(pdf_value) * 100) if pdf_value not in (None, 0.0) else None
                    is_large_amount = (pdf_value and abs(pdf_value) > 100_000)
                    
                    if delta < 0.01:
                        status = "OK"
                    elif is_large_amount and delta_pct is not None and delta_pct < 1.5:
                        status = "OK"
                    elif abs(result_value * 100 - pdf_value) < max(abs(pdf_value) * 0.05, 1.0) or abs(result_value - pdf_value * 100) < max(abs(pdf_value * 100) * 0.05, 1.0):
                        status = "ANOMALIE UNITE PROBABLE (facteur ~100)"
                    else:
                        status = "ECART SIGNIFICATIF"

            mapped_status = _COMPARISON_STATUSES.get(status, "mismatched")
            matched += 1 if mapped_status == "matched" else 0
            mismatched += 1 if mapped_status == "mismatched" else 0
            pdf_only += 1 if mapped_status == "pdf_only" else 0
            csv_only += 1 if mapped_status == "csv_only" else 0

            rows.append({
                "reference": indicator,
                "date": str(row.get("Source PDF", "")),
                "amount": result_value if not pd.isna(result_value) else pdf_value,
                "status": mapped_status,
                "pdf_amount": pdf_value,
                "csv_amount": result_value,
                "difference": None if pd.isna(pdf_value) or pd.isna(result_value) else round(abs(result_value - pdf_value), 6),
            })

        total_comparable = matched + mismatched
        score = (matched / total_comparable * 100) if total_comparable > 0 else 0.0
        return {
            "matched": matched,
            "mismatched": mismatched,
            "pdf_only": pdf_only,
            "csv_only": csv_only,
            "score": round(score, 2),
            "details": rows,
            "pdf_only_details": [row for row in rows if row["status"] == "pdf_only"],
            "csv_only_details": [row for row in rows if row["status"] == "csv_only"],
            "transactions": rows,
        }

    if csv_df.columns.size >= 2 and _normalize_key(csv_df.columns[0]) in {"métrique", "metrique", "metric"}:
        bank_column = _pick_bank_column(csv_df, pdf_name=pdf_name)
        if bank_column is None:
            return {
                "matched": 0,
                "mismatched": 0,
                "pdf_only": 0,
                "csv_only": 0,
                "score": 0.0,
                "details": [],
                "pdf_only_details": [],
                "csv_only_details": [],
                "transactions": [],
                "error": "Could not determine a bank column in the wide workbook.",
            }

        pdf_map = {str(item.get("Indicateur", "")).strip().lower(): item for item in pdf_transactions}
        rows = []
        matched = 0
        mismatched = 0
        lei = _lei_for_col(bank_column)

        for _, row in csv_df.iterrows():
            indicator = str(row.iloc[0]).strip()
            if not indicator or indicator.lower() == "nan":
                continue

            result_cell = row.get(bank_column)
            if result_cell is None or (isinstance(result_cell, float) and pd.isna(result_cell)) or str(result_cell).strip() == "":
                result_value = None
            else:
                result_value = _parse_amount(result_cell)

            pdf_row = pdf_map.get(indicator.lower())
            pdf_value = pdf_row.get("Valeur PDF (EBA)") if pdf_row else None
            source_pdf = pdf_row.get("Source PDF") if pdf_row else ""

            if pdf_row is None and result_value is None:
                status = "PDF non disponible (pas de LEI)"
                pdf_value = None
            elif pdf_row is None:
                status = "Non trouvé dans le PDF"
                csv_only += 1
                pdf_value = None
            elif result_value is None:
                status = "Absent du fichier resultats"
                pdf_only += 1
                pdf_value = pdf_row.get("Valeur PDF (EBA)") if pdf_row else None
            else:
                delta = abs(result_value - pdf_value)
                delta_pct = (delta / abs(pdf_value) * 100) if pdf_value not in (None, 0) else None
                is_large_amount = (pdf_value and abs(pdf_value) > 100_000)
                
                if delta < 0.0005 or (pdf_value == 0 and result_value == 0):
                    status = "OK"
                    matched += 1
                elif is_large_amount and delta_pct is not None and delta_pct < 1.5:
                    status = "OK"
                    matched += 1
                elif pdf_value not in (None, 0) and abs(result_value * 100 - pdf_value) < max(abs(pdf_value) * 0.05, 1.0):
                    status = "ANOMALIE UNITE PROBABLE (facteur ~100)"
                    mismatched += 1
                else:
                    status = "ECART SIGNIFICATIF"
                    mismatched += 1

            if pdf_row is None or result_value is None:
                pdf_value = None
                delta = None
                delta_pct = None
            else:
                delta = None if pdf_value is None or result_value is None else abs(result_value - pdf_value)
                delta_pct = None if delta is None or pdf_value in (None, 0) else (delta / abs(pdf_value) * 100)

            rows.append({
                "Entité": _output_name_for_col(bank_column),
                "LEI": lei,
                "Indicateur": indicator,
                "Valeur resultats": result_value,
                "Valeur PDF (EBA)": pdf_value,
                "Ecart": delta,
                "Ecart %": delta_pct,
                "Statut": status,
                "Source PDF": source_pdf,
                "reference": indicator,
                "date": source_pdf,
                "amount": result_value if not pd.isna(result_value) else (pdf_value or 0.0),
                "status": "matched" if status == "OK" else "ecart",
                "pdf_amount": pdf_value,
                "csv_amount": result_value,
                "difference": delta,
            })

        total_comparable = matched + mismatched
        score = (matched / total_comparable * 100) if total_comparable > 0 else 0.0
        return {
            "matched": matched,
            "mismatched": mismatched,
            "pdf_only": pdf_only,
            "csv_only": csv_only,
            "score": round(score, 2),
            "details": rows,
            "pdf_only_details": [],
            "csv_only_details": [],
            "transactions": rows,
            "bank_column": bank_column,
        }

    if not pdf_transactions:
        return {
            "matched": 0, "mismatched": 0, "pdf_only": 0, "csv_only": 0,
            "score": 0.0, "details": [], "pdf_only_details": [], "csv_only_details": []
        }

    ref_col = _find_column(csv_df, _REF_ALIASES)
    date_col = _find_column(csv_df, _DATE_ALIASES)
    
    debit_col = _find_column(csv_df, ["débit", "debit"])
    credit_col = _find_column(csv_df, ["crédit", "credit"])
    amt_col = _find_column(csv_df, ["amount", "montant"])

    if not ref_col or (not amt_col and not debit_col and not credit_col):
        return {
            "matched": 0, "mismatched": 0, "pdf_only": 0, "csv_only": 0,
            "score": 0.0, "details": [], "pdf_only_details": [], "csv_only_details": [],
            "error": f"Could not determine required columns in CSV. Found: {list(csv_df.columns)}"
        }

    csv_records = {}
    for idx, row in csv_df.iterrows():
        raw_ref = str(row[ref_col]).strip()
        ref_key = raw_ref.lower()
        
        if debit_col or credit_col:
            deb_val = _parse_amount(row[debit_col]) if debit_col else 0.0
            cred_val = _parse_amount(row[credit_col]) if credit_col else 0.0
            amt_val = deb_val if deb_val != 0.0 else cred_val
        else:
            amt_val = abs(_parse_amount(row[amt_col]))

        # Skip rows without reference labels or total rows
        if raw_ref == "" or pd.isna(row[ref_col]) or "total" in ref_key:
            continue

        if ref_key not in csv_records:
            csv_records[ref_key] = []
            
        csv_records[ref_key].append({
            "index": idx,
            "reference": raw_ref,
            "amount": amt_val,
            "date": str(row[date_col]).strip() if date_col else "",
            "matched": False
        })

    matched = []
    mismatched = []
    pdf_only = []

    for tx in pdf_transactions:
        pdf_ref = tx.get("Reference", "")
        ref_key = str(pdf_ref).strip().lower()
        pdf_amount = tx.get("Amount", 0.0)

        # Ignore non-transaction headers from PDF outputs
        if "total" in ref_key or "solde" in ref_key:
            continue

        candidates = csv_records.get(ref_key, [])
        unmatched_candidate = next((c for c in candidates if not c["matched"]), None)

        if unmatched_candidate is None:
            pdf_only.append({
                "reference": pdf_ref,
                "date": tx.get("Date", ""),
                "amount": pdf_amount
            })
            continue

        unmatched_candidate["matched"] = True
        csv_amount = unmatched_candidate["amount"]

        if abs(pdf_amount - csv_amount) < 0.01:
            matched.append(tx)
        else:
            mismatched.append({
                "reference": pdf_ref,
                "date": tx.get("Date", ""),
                "pdf_amount": pdf_amount,
                "csv_amount": csv_amount,
                "difference": round(abs(pdf_amount - csv_amount), 2)
            })

    csv_only = []
    for ref_key, items in csv_records.items():
        for item in items:
            # Only count as csv-only if it was not matched and isn't a descriptive header
            if not item["matched"] and "solde" not in ref_key:
                csv_only.append({
                    "reference": item["reference"],
                    "date": item["date"],
                    "amount": item["amount"]
                })

    total_pdf_txs = len(pdf_transactions)
    score = (len(matched) / total_pdf_txs * 100) if total_pdf_txs > 0 else 0.0

    return {
        "matched": len(matched),
        "mismatched": len(mismatched),
        "pdf_only": len(pdf_only),
        "csv_only": len(csv_only),
        "score": round(score, 2),
        "details": mismatched,
        "pdf_only_details": pdf_only,
        "csv_only_details": csv_only
    }